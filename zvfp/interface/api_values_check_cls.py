# ！/usr/bin/env python
# -*-coding: utf-8 -*-

import requests
from openpyxl import load_workbook
from req_data import *



# 定义一个接口字段测试比对类
class ApiFieldValuesCheck():

    def __init__(self, excel_path,excel_sheet_name):
        '''
        :param url: 请求url
        :param body: 请求报文
        :param headers: 请求头
        :param excel_path: 测试数据excel文件路径
        :param excel_sheet_name: excel文件sheet页名称
        '''
        self.excel_path = excel_path
        self.excel_sheet_name = excel_sheet_name
        self.field_name_list = [] #接收接口字段名称列表
        self.field_value_list = []  #接收接口字段值的列表

    def json_data_to_list(self, json_data):
        '''
        将响应报文拆解成字段及其值存放到列表
        :param json_data: 响应报文的json串
        :return: field_name_list(响应报文字段列表),field_name_list(响应报文字段值)
        '''

        for key, value in json_data.items():
            #判断字段值是否为字典，如果为字典，将该值作为新的参数，调用本函数
            if isinstance(value, dict):
                self.json_data_to_list(value)
            #若字段值不是字典，判断字段值是否为列表
            elif isinstance(value, list):
                if len(value) != 0:
                    #循环列表中的元素
                    for list_value in value:
                        #若循环列表中的元素为字典，将该值作为新的参数，调用本函数
                        if isinstance(list_value, dict):
                            self.json_data_to_list(list_value)
                #若列表为空，则直接将字段名称和None写入对应的列表中
                else:
                    self.field_name_list.append(key)
                    self.field_value_list.append(None)
            #若字段值非字典且非列表，则将字段名称和字段值写入对应的列表中
            else:
                self.field_name_list.append(key)
                self.field_value_list.append(value)
        return self.field_name_list, self.field_value_list

    def loop_write_col(self, start_col, list_name):
        '''
        将列表数据循环写入excel某列

        :param start_col: 需要写入的列数，int类型
        :param list_name: 列表名称
        :param save_excel_name: 保存的excel名称
        :return:
        '''
        wb = load_workbook(self.excel_path)
        ws = wb[self.excel_sheet_name]
        row_count = 0
        while row_count < len(list_name):
            #  row=row_count + 2, 跳过表头，从第二行开始写
            ws.cell(row=row_count+2, column=start_col,value=list_name[row_count]).value
            row_count += 1
        wb.save(self.excel_path)




    def read_excel_data(self):
        '''
        读取excel测试数据，保存到列表
        :param excel_path: excel测试数据保存路径
        :param excel_sheet_name: excel接口测试数据所在的sheet页名称
        :return:
        '''
        wb = load_workbook(self.excel_path)
        # ws = wb["risk_control"]  # "risk_control"
        ws = wb[self.excel_sheet_name]  # excel sheet 页名称，可以将多个接口的测试数据维护在一个excel表里

        # 读取excel文件数据，保存到列表
        excel_data = []  # 列表套列表，相当于二维数组
        for row in ws.iter_rows(min_col=1, min_row=2, max_row=ws.max_row, max_col=ws.max_column):  #
            single_row = []
            for cell in row:
                single_row.append(cell.value)
            excel_data.append(single_row)

        return excel_data



    def api_test_result(self, excel_data):
        '''
        将接口实际返回值与预期值比对，得出比对结果
        :param excel_data: 接口的excel数据（api_values_check.xlsx）
        :return: 接口字段返回实际值与预期值的比对结果列表
        '''
        excel_data_index = 0
        test_api_result_list = []  # 保存测试结果列表
        while excel_data_index < len(excel_data):
            if excel_data[excel_data_index][2] == "是" :
                if excel_data[excel_data_index][1] == excel_data[excel_data_index][3]:
                    test_api_result_list.append("通过")
                else:
                    test_api_result_list.append("失败")
            else:
                test_api_result_list.append("不需要比对")
                # pass
            excel_data_index += 1

        return test_api_result_list


    def api_field_test(self,req_data):
        '''
        接口请求响应报文字段内如与预期值比对
            响应报文字段存入excel模版并与逾期值比对
            比对结果自动填充
        :return:
        '''

        # 第二步：响应报文字段与值拆解，存入列表
        api_field_names, api_field_values = self.json_data_to_list(req_data)

        # 第三步：将字段及其值先入excel模版
        # 在模版第1列写入系统返回字段
        self.loop_write_col(1, api_field_names)
        # 在模版第2列写入字段值
        self.loop_write_col(2, api_field_values)

        # 第四步：读取excel数据，为比对做数据导入
        api_excel_data = self.read_excel_data()


        # 第五步：接口实际值与预期值比对
        test_api_result_list = self.api_test_result(api_excel_data)

        # 第六步： 将测试结果写入excel
        # 在模版第5列写入测试比结果
        self.loop_write_col(5, test_api_result_list)

if __name__ == '__main__':
    test_obj = ApiFieldValuesCheck(url, body, headers, excel_path, risk_sheet_name)
    test_obj.api_field_test()