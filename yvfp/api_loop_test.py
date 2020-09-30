# ！/usr/bin/env python
# -*-coding: utf-8 -*-
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import json

# 定义一个API类
class ApiMethod():

    def __init__(self, url, data, header=None):
        self.url = url
        self.data = data
        self.header = header

    # post请求
    def post_request(self):
        if self.header != None:
            print("请求报文：")
            print(self.data)
            res = requests.post(url=self.url, json=self.data, headers=self.header, verify=False)
            print("响应码：", res)
            print("响应报文：")
            print(res.text)
        else:
            print("请求报文：")
            print(self.data)
            res = requests.post(url=self.url, json=self.data, verify=False)
            print("响应码：", res)
            print("响应报文：")
            print(res.text)
        return res.json()

    # post 返回
    def post_respond(self):
        post_res = self.post_request()
        print(post_res)
        return post_res

    def get_request(self):
        if self.header != None:
            res = requests.get(url=self.url, json=self.data, headers=self.header, verify=False)
        else:
            res = requests.get(url=self.url, json=self.data, verify=False)
        return res.json()

    def get_respond(self):
        get_res = self.get_request()
        return get_res


excel_path = r"D:\soft\test_api_data.xlsx"  # excel测试数据文件所在路径

def read_excel_data(excel_path,excel_sheet_name):
    '''
    读取excel测试数据，保存到列表
    :param excel_path: excel测试数据保存路径
    :param excel_sheet_name: excel接口测试数据所在的sheet页名称
    :return:
    '''
    wb = load_workbook(excel_path)


    # ws = wb["risk_control"]  # "risk_control"
    ws = wb[excel_sheet_name]  # excel sheet 页名称，可以将多个接口的测试数据维护在一个excel表里

    # 读取excel文件数据，保存到列表
    excel_data = []  # 列表套列表，相当于二维数组
    for row in ws.iter_rows(min_col=1, min_row=2, max_row=ws.max_row, max_col=ws.max_column):  #
        single_row = []
        for cell in row:
            single_row.append(cell.value)
        excel_data.append(single_row)
    print(excel_data)

    return excel_data

risk_sheet = "risk_control"
vfp_sheet = "vfp_pre_approve"

def api_multiple_data():
    '''
    读取excel多组测试数据，对单接口进行多组数据的接口测试
    :return:
    '''
    excel_data_list = read_excel_data(excel_path,risk_sheet)  # 调excel函数，获取测试多组测试数据


    # print(sys._getframe().f_code.co_name)  # 获取当前函数名称
    with open('%s.txt' %sys._getframe().f_code.co_name,'w',encoding='utf-8') as wf:
        wf.write("")
        wf.write("读取到excel的所有入参如下数据：" )
        wf.write("\n")
        json.dump(excel_data_list, wf, ensure_ascii=False)
        wf.write("\n")
        wf.write("黄金分割线".center(110, "*"))
        wf.write("\n")

    excel_data_index = 0
    while excel_data_index < len(excel_data_list):
        print("第%s组请求数据：%s" %(int(excel_data_index+1),excel_data_list[excel_data_index]))  # 为打印显示用的，可以不用管
        # print(excel_data_list[excel_data_index])
        # print(excel_data_list[excel_data_index][0])  # excel第一列（身份证）
        # print(excel_data_list[excel_data_index][1])  # excel第二列（姓名）
        # print(excel_data_list[excel_data_index][2])  # excel第二列（手机号）
        # 请求接口数据，请求地址，请求体可按照格式修改
        url = "http://10.138.60.149:9056/api/PolicyDecision/Execute"
        headers = {"Content-Type": "application/json"}
        # CustName = "汪雪兰"
        # identityNo = 411624199105121215
        # mobile = 13510245116
        body = {"flowID": "0556227395dd425d917fbdbf8998814f",
                "flowVersion": -1.0,
                "identityNo": excel_data_list[excel_data_index][0],
                "inputVbs": "{\"SEX\":\"男\",\"AGE\":29,\"region_online\":\"陕西省宝鸡市风翔县龙潭江街8座_上海市\",\"IsHousehold_Online\":\"否\",\"CustomerSource\":\"2\",\"Front_Education\":\"硕士及以上\",\"Caller\":\"豆豆钱\",\"Call_Type\":\"授信\",\"Nation\":\"-1\",\"ChannelName\":\"奇虎360\",\"ChannelType\":\"DSP\",\"RiskCustomerFlow\":\"-1\",\"isGroupSaleBootCustomer\":\"否\"}",
                "dynamicParams": "{\"App_Name\":\"doudouhua_and\",\"BusType\":\"DOUDOUQIAN\",\"CreditReportSn\":\"\",\"CustName\":\"%s\",\"PerAddress\":\"陕西省宝鸡市风翔县龙潭江街8座\",\"ProcessId\":\"468122184137854976\",\"ReportId\":\"\",\"ReportType\":\"\",\"identityNo\":\"%s\",\"mobile\":\"%s\",\"mobile_id\":\"-999\",\"name\":\"汪雪兰\",\"product_id\":53,\"report_id\":\"\",\"tongdunToken\":\"-1\",\"registerMobile\":\"13510245116\",\"userId\":\"dd_300022702354\",\"vcreditid\":\"-1\",\"os_name\":\"-1\"}" % (
                excel_data_list[excel_data_index][1], excel_data_list[excel_data_index][0], excel_data_list[excel_data_index][2]),
                "callFrom": "豆豆钱",
                "flowName": None,
                "engineType": "VRC",
                "isDebug": 1}

        api_request = ApiMethod(url, body, headers)
        req_data = api_request.post_request()
        print(req_data)  # 非必须，方便查看

        # 接口响应报文写入文件，方便查看响应报文，暂不要求掌握
        with open('%s.txt' %sys._getframe().f_code.co_name, 'a+', encoding='utf-8') as wf:

            wf.write("第%s组请求数据：%s" %(int(excel_data_index+1),excel_data_list[excel_data_index]))
            wf.write("\n")
            json.dump(req_data,wf,ensure_ascii=False)
            wf.write("\n")
            wf.write("黄金分割线".center(110,"*"))
            wf.write("\n")
            wf.flush()

        print("黄金分割线".center(110,"*"))

        excel_data_index +=1

# read_excel_data(excel_path)
# api_multiple_data()


def vfp_multiple_data():
    '''
    读取excel多组测试数据，对单接口进行多组数据的接口测试
    :return:
    '''
    excel_data_list = read_excel_data(excel_path, vfp_sheet)  # 调excel函数，获取测试多组测试数据
    # print("**************************")
    # print(sys._getframe().f_code.co_name)  # 获取当前函数名称
    with open('%s.txt' %sys._getframe().f_code.co_name,'w',encoding='utf-8') as wf:
        wf.write("")
        wf.write("读取到excel的所有入参如下数据：" )
        wf.write("\n")
        json.dump(excel_data_list, wf, ensure_ascii=False)
        wf.write("\n")
        wf.write("黄金分割线".center(110, "*"))
        wf.write("\n")

    excel_data_index = 0  #初始化列表索引，为循环设置起始值
    while excel_data_index < len(excel_data_list):
        print("第%s组请求数据：%s" %(int(excel_data_index+1),excel_data_list[excel_data_index]))  # 为打印显示用的，可以不用管

        headers = {"Content-Type": "application/json"}
        headers1 = {"Content-Type": "application/json",
                    "apiToken":"e6475549-4681-4b12-9ca0-d28c0e03230a",
                    "sync":"false",
                    "X-CLIENT-ID":"123"
                    }
        init_url = "http://api-gateway-t2.dev.vcredit.com.local/fund/integratedQuery/"
        init_body = {
        "serviceCode": "init",
        "idCardNo": "653223197304301488",
        "customerName": "许桂芳",
        "idType": "20",
        "pid": "319174329378295808",
        "productCode": excel_data_list[excel_data_index][4],
        "tractionId": "zlmtestchushihua1595466873",
        "routeAuthToken": None
        }

        api_request = ApiMethod(init_url, init_body, headers1)
        req_data = api_request.post_request()
        routeAuthToken = req_data['data']['routeAuthToken']  # 获取token
        # print(req_data)  # 非必须，方便查看

        # 请求接口数据，请求地址，请求体可按照格式修改
        pre_approve_url = "http://api-gateway-t2.dev.vcredit.com.local/fund/routeDataSubmit"

        # CustName = "汪雪兰"
        # identityNo = 411624199105121215
        # mobile = 13510245116
        pre_approve_body = {
        "transactionId": "zlmtestpre1595466873",
        "productCode": excel_data_list[excel_data_index][4],
        "routeAuthToken": routeAuthToken,
        "personalBasicInfo": {
        "idType": "CETP000001",
        "idNo": "653223197304301488",
        "idValidStart": "20191212",
        "idValidEnd": "99990101",
        "customerName": "许桂芳",
        "birthdate": "19740101",
        "customerType": "CTTP000003",
        "email": "965015785@qq.com",
        "sex": "F",
        "nation": "汉",
        "marriage": "MAST000002",
        "education": "EDBK000003",
        "degree": "DEGR000006",
        "phoneNo": "15674363264",
        "customerSource": "线上",
        "address": "广东省兴宁市罗岗镇徐坑村丑子7号",
        "postAddr": "广东省兴宁市罗岗镇徐坑村丑子7号",
        "isOldCustomer": "0",
        "isMemberFeeExpired": "0",
        "loanSpecialUserType": "2",
        "channelName": "fhd",
        "liveVerificationResult": "0",
        "deviceNo": "865371039361513",
        "idIssuingAuthority": "上海市公安局杨浦分局",
        "registerTime": "20171221120000",
        "channelType": "-1"
        },
        "loanApplicationInfo": {
        "applicantType": "APTP000001",
        "applyArea": "上海市",
        "applDate": "2019-01-01",
        "loanType": "ORTP000001",
        "riskManLoanType": "ROTP000001",
        "creditUpfrontLabel": "-1",
        "lastSettlementDate": "-",
        "loanCurrency": "CURY000001",
        "applyAmount": excel_data_list[excel_data_index][0],
        "applyPeriod": excel_data_list[excel_data_index][1],
        "changePeriod": "3",
        "inRate": "0.105",
        "monthlyInterestRate": "0.008500",
        "monthlyServiceRate": "0.008500",
        "highMonthlyRate": "0.02",
        "specialMonthlyRate": "0.02",
        "formalitiesRate": "0",
        "formalitiesFee": "0",
        "deductionAmount": "0",
        "creditChannel": "CRCH000002",
        "creditInfoCacheFlag": "1",
        "creditCheckTime": "20171221145658",
        "creditAmt": "50000.000000",
        "customerLevel": "3",
        "creditReportId": "1164018586902069249",
        "firstLoanFlag": "1",
        "supplementCustomerType": excel_data_list[excel_data_index][2],
        "mobileReportId": "jxl_19802539",
        "repayMethod": "RPMD000001",
        "dueDayOpt": None,
        "purposeArea": "LUCS000009",
        "purpose": "LUTP000006",
        "customerMark": excel_data_list[excel_data_index][3],
        "repayType": "RPMD000001"
        },
        "personalJobInfo": {
        "incomeMonth": "5000",
        "incomeOfYear": "3",
        "occupation": "OCCU000004",
        "accfundstate": "2",
        "industry": "INDU000021",
        "companyName": "上海市智慧金融有限公司",
        "companyAddr": "上海市虹口区四川北路",
        "indivEmpTyp": "其他",
        "preSubjectLabel": "7",
        "position": "POSI000008",
        "positionalTitle": "POTI000005",
        "indivEmpYrs": "10",
        "indivEmpTel": "3131"
        },
        "bankCardInfo": {
        "bankNo": "105100000017",
        "bankName": "建设银行",
        "bankType": "BCTP000001",
        "bankCardNo": "6217000416301085017",
        "bankBranchName": "建设银行",
        "bindPhone": "15674363264",
        "repayBankNo": "105100000017",
        "repayBankCardNo": "6217000416301085017",
        "repayBindPhone": "15674363264",
        "repayBankName": "建设银行",
        "applAcTyp": "AUTP000001",
        "applIdTyp": "CETP000001",
        "applAcKind": "ACTP000001",
        "applAcNam": "张一山",
        "applIdNo": "653223197304301488",
        "repayBankType": "BCTP000001",
        "isFourCheck": "1",
        "fourCheckResult": "1"
        },
        "familyInfo": {
        "homeArea": "广东省兴宁市罗岗镇徐坑村丑子7号",
        "localResident": "HRTP000001",
        "homeInfo": "HSST000013",
        "familyIncomeMonth": "10000",
        "hasChildren": "1",
        "hasCar": "0",
        "hasCarCredit": "1",
        "familyTel": "400211211",
        "familyZip": "412300",
        "housePrice": "100000"
        },
        "contactsInfo": [
        {
            "contactName": "杨紫",
            "contactRel": "RELA000001",
            "contactMobile": "15514560001",
            "contactIdType": "CETP000001",
            "contactIdNo": "110101197401011894"
        },
        {
            "contactName": "陈涛",
            "contactRel": "RELA000002",
            "contactMobile": "15514560002",
            "contactIdType": "CETP000001",
            "contactIdNo": "110101197401011915"
        }
        ],
        "fundSpecialInfo": {
        "referType": "04"
        },
        "idCardImageInfo": {
        "idFrontImage": "http://10.139.60.163/group1/M01/60/18/Cos8o11cspSAFod8AACO72pQ8oY140.jpg",
        "idBackImage": "http://10.139.60.163/group1/M01/60/18/Cos8o11csmmAZzJFAAR33xq9DTI426.jpg",
        "livingBodyImage": "http://10.139.60.163/group1/M01/60/18/Cos8o11csr-AFJEbAACXHvIDsGQ282.jpg",
        "signPic": "http://10.139.60.163/group1/M01/60/1A/Cos8o11dF6qAKL-oAABi5feQHFE174.png",
        "signType": None
        },
        "frontControlInfo": {
        "noSettleFundCode": [],
        "maxAuditTimes": "10"
        }
        }

        api_request = ApiMethod(pre_approve_url, pre_approve_body, headers1)
        req_data = api_request.post_request()
        # print(pre_approve_body)
        print(req_data)  # 非必须，方便查看

        # 接口响应报文写入文件，方便查看响应报文，暂不要求掌握
        with open('%s.txt' %sys._getframe().f_code.co_name, 'a+', encoding='utf-8') as wf:

            wf.write("第%s组请求数据：%s" %(int(excel_data_index+1),excel_data_list[excel_data_index]))
            wf.write("\n")
            json.dump(req_data,wf,ensure_ascii=False)
            wf.write("\n")
            wf.write("黄金分割线".center(110,"*"))
            wf.write("\n")
            wf.flush()

        print("黄金分割线".center(110,"*"))

        excel_data_index +=1

def call_api():
    '''
    调用单个或者多个api流程
    :return:
    '''
    # read_excel_data(excel_path,risk_sheet)  # 读excel数据
    # api_multiple_data()

    read_excel_data(excel_path, vfp_sheet)  # 读excel数据
    vfp_multiple_data()  # 循环调单接口，实现单接口多数据测试

call_api()
# if __name__ == '__main__':
#         call_api()