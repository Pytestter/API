# ！/usr/bin/env python
# -*-coding: utf-8 -*-

import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from req_data import *

# 定义一个API类
class ApiMethod():

    def __init__(self, url, data, header=None):
        self.url = url
        self.data = data
        self.header = header

    # post请求
    def post_request(self):
        if self.header != None:
            print("请求报文：")
            print(self.data)
            res = requests.post(url=self.url, json=self.data, headers=self.header, verify=False)
            print("响应码：", res)
            print("响应报文：")
            print(res.text)
        else:
            print("请求报文：")
            print(self.data)
            res = requests.post(url=self.url, json=self.data, verify=False)
            print("响应码：", res)
            print("响应报文：")
            print(res.text)
        return res.json()

    # post 返回
    def post_respond(self):
        post_res = self.post_request()
        print(post_res)
        return post_res

    def get_request(self):
        if self.header != None:
            res = requests.get(url=self.url, json=self.data, headers=self.header, verify=False)
        else:
            res = requests.get(url=self.url, json=self.data, verify=False)
        return res.json()

    def get_respond(self):
        get_res = self.get_request()
        return get_res







def json_data_to_list(json_data):
    '''
    将响应报文拆解成字段及其值存放到列表
    :param json_data: 响应报文的json串
    :return: field_name_list(响应报文字段列表),field_name_list(响应报文字段值)
    '''

    field_name_list = []  # 接收接口字段名称列表
    field_value_list = []  # 接收接口字段值的列表

    #  循环遍历，提取字段名称及其值，添加到列表
    for k1,v1 in json_data.items():
        # 如果是字典嵌套字典，遍历字典嵌套，目前实现三层
        if isinstance(v1,dict) and v1:
            for k2,v2 in v1.items():
                if isinstance(v2, dict) and v2:
                    for k3,v3 in v2.items():
                        field_name_list.append(k3)
                        field_value_list.append(v3)
                        # print(k3,v3)
                else:
                    field_name_list.append(k2)
                    field_value_list.append(v2)
                    # print(k2,v2)

        # 如果是字典嵌套列表，遍历列表嵌套，目前实现三层
        elif isinstance(v1,list) and v1:
            i = 0
            while i < len(v1):
                for k2, v2 in v1[i].items():
                    if isinstance(v2, list) and v2:
                        j = 0
                        while j < len(v2):
                            for k3, v3 in v2[j].items():
                                field_name_list.append(k3)
                                field_value_list.append(v3)
                                # print(k3, v3)
                            j += 1
                    else:
                        field_name_list.append(k2)
                        if v2:
                            field_value_list.append(v2)
                        else:
                            field_value_list.append(None)
                        # print(k2,v2)
                    i += 1

        else:
            field_name_list.append(k1)
            if v1 :
                field_value_list.append(v1)
            else:
                field_value_list.append(None)
            # print(k1,v1)
    print("响应报文字段名称列表：",field_name_list)
    print("响应报文字段值列表：",field_value_list)

    return  field_name_list, field_value_list





# print("********************************************************")
# json_data_to_list(r1)



def loop_write_col(start_col,list_name,save_excel_name, excel_sheet_name ):
    '''
    将列表数据循环写入excel某列

    :param start_col: 需要写入的列数，int类型
    :param list_name: 列表名称
    :param save_excel_name: 保存的excel名称
    :return:
    '''
    wb = load_workbook(excel_path)
    ws = wb[excel_sheet_name]
    row_count = 0
    while row_count < len(list_name):
        #  row=row_count + 2, 跳过表头，从第二行开始写
        ws.cell(row=row_count+2, column=start_col,value=list_name[row_count]).value
        row_count += 1
    wb.save(save_excel_name)




def read_excel_data(excel_path,excel_sheet_name):
    '''
    读取excel测试数据，保存到列表
    :param excel_path: excel测试数据保存路径
    :param excel_sheet_name: excel接口测试数据所在的sheet页名称
    :return:
    '''
    wb = load_workbook(excel_path)
    # ws = wb["risk_control"]  # "risk_control"
    ws = wb[excel_sheet_name]  # excel sheet 页名称，可以将多个接口的测试数据维护在一个excel表里

    # 读取excel文件数据，保存到列表
    excel_data = []  # 列表套列表，相当于二维数组
    for row in ws.iter_rows(min_col=1, min_row=2, max_row=ws.max_row, max_col=ws.max_column):  #
        single_row = []
        for cell in row:
            single_row.append(cell.value)
        excel_data.append(single_row)
    print(excel_data)

    return excel_data



def api_test_result(excel_data):
    '''
    将接口实际返回值与预期值比对，得出比对结果
    :param excel_data: 接口的excel数据（api_values_check.xlsx）
    :return: 接口字段返回实际值与预期值的比对结果列表
    '''
    excel_data_index = 0
    test_api_result_list = []  # 保存测试结果列表
    while excel_data_index < len(excel_data):
        if excel_data[excel_data_index][2] == "是" :
            if excel_data[excel_data_index][1] == excel_data[excel_data_index][3]:
                test_api_result_list.append("通过")
            else:
                test_api_result_list.append("失败")
        else:
            test_api_result_list.append("不需要比对")
            # pass
        excel_data_index += 1

    return test_api_result_list


def api_field_test():
    '''
    接口请求响应报文字段内如与预期值比对
        响应报文字段存入excel模版并与逾期值比对
        比对结果自动填充
    :return:
    '''

    # 第一步：调接口获取响应报文
    api_obj = ApiMethod(url, body, headers)
    req_data = api_obj.post_request()

    # 第二步：响应报文字段与值拆解，存入列表
    api_field_names, api_field_values = json_data_to_list(req_data)

    # 第三步：将字段及其值先入excel模版
    # 在模版第1列写入系统返回字段
    write_field_name = loop_write_col(1, api_field_names, excel_path, risk_sheet_name)
    # 在模版第2列写入字段值
    write_field_values = loop_write_col(2, api_field_values, excel_path, risk_sheet_name)

    # 第四步：读取excel数据，为比对做数据导入
    api_excel_data = read_excel_data(excel_path, risk_sheet_name)  # risk_sheet_name 为excel名称
    # print(api_excel_data)
    # print(len(vfp_excel_data))
    # print(api_excel_data[0][1])  # 第二列接口返回值
    # print(api_excel_data[0][2])  # 第三列，是否需要比对
    # print(api_excel_data[0][3])  # 第四列，预期值

    # 第五步：接口实际值与预期值比对
    test_api_result_list = api_test_result(api_excel_data)

    # 第六步： 将测试结果写入excel
    # 在模版第5列写入测试比结果
    write_field_is_filled = loop_write_col(5, test_api_result_list, excel_path, risk_sheet_name)

if __name__ == '__main__':
    api_field_test()